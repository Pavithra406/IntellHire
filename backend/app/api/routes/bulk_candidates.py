import io
import openpyxl
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.core.database import get_db
from app.core.security import require_hr
from app.repositories.user_repository import UserRepository
from app.repositories.candidate_repository import CandidateRepository
from app.repositories.job_repository import JobRepository

router = APIRouter(prefix="/candidates/bulk", tags=["Bulk Candidates"])


@router.get("/template")
def download_template(hr=Depends(require_hr)):
    """Download Excel template for bulk candidate upload."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Candidates"

    # Header row
    headers = ["full_name", "email", "password", "job_id"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="6D28D9")
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")

    # Sample rows
    samples = [
        ["Arjun Sharma", "arjun@example.com", "pass1234", 1],
        ["Priya Patel", "priya@example.com", "pass5678", 1],
        ["Rahul Verma", "rahul@example.com", "pass9012", 1],
    ]
    for row_data in samples:
        ws.append(row_data)

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 10

    # Add instructions sheet
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        ["Column", "Description", "Required"],
        ["full_name", "Candidate's full name", "Yes"],
        ["email", "Unique email address for login", "Yes"],
        ["password", "Login password (min 6 chars)", "Yes"],
        ["job_id", "Job Opening ID from the HR portal", "Yes"],
        ["", "", ""],
        ["Note:", "Do not change column headers", ""],
        ["Note:", "Each email must be unique", ""],
        ["Note:", "Get job_id from HR portal → Jobs page", ""],
    ]
    for row in instructions:
        ws2.append(row)
    ws2.column_dimensions["A"].width = 15
    ws2.column_dimensions["B"].width = 45
    ws2.column_dimensions["C"].width = 10

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=candidate_upload_template.xlsx"}
    )


@router.post("/upload")
async def bulk_upload(
    file: UploadFile = File(...),
    job_id: int = None,
    db: Session = Depends(get_db),
    hr=Depends(require_hr)
):
    """Upload Excel file to bulk create candidate accounts."""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only .xlsx or .xls files are accepted")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
    except Exception:
        raise HTTPException(400, "Could not read Excel file. Use the provided template.")

    # Read headers from row 1
    headers = [str(ws.cell(1, c).value).strip().lower() if ws.cell(1, c).value else ""
               for c in range(1, ws.max_column + 1)]

    required = {"full_name", "email", "password", "job_id"}
    missing = required - set(headers)
    if missing:
        raise HTTPException(400, f"Missing columns: {', '.join(missing)}. Download the template.")

    col = {h: i for i, h in enumerate(headers)}

    user_repo = UserRepository(db)
    candidate_repo = CandidateRepository(db)
    job_repo = JobRepository(db)

    results = []
    errors = []

    for row_num in range(2, ws.max_row + 1):
        def cell(key):
            v = ws.cell(row_num, col[key] + 1).value
            return str(v).strip() if v is not None else ""

        full_name = cell("full_name")
        email = cell("email")
        password = cell("password")
        jid = int(cell("job_id")) if cell("job_id").isdigit() else job_id

        # Skip empty rows
        if not full_name and not email:
            continue

        # Validate row
        row_errors = []
        if not full_name:
            row_errors.append("full_name is empty")
        if not email or "@" not in email:
            row_errors.append("invalid email")
        if not password or len(password) < 6:
            row_errors.append("password too short (min 6 chars)")
        if not jid:
            row_errors.append("job_id missing")

        if row_errors:
            errors.append({"row": row_num, "email": email, "errors": row_errors})
            continue

        # Check job belongs to HR
        job = job_repo.get_by_id(jid)
        if not job or job.hr_user_id != hr.id:
            errors.append({"row": row_num, "email": email, "errors": [f"Job ID {jid} not found or not yours"]})
            continue

        # Check duplicate email
        if user_repo.get_by_email(email):
            errors.append({"row": row_num, "email": email, "errors": ["Email already registered"]})
            continue

        try:
            user = user_repo.create(email, full_name, password, role="candidate")
            candidate_repo.create(user.id, jid, hr.id)
            results.append({
                "row": row_num,
                "full_name": full_name,
                "email": email,
                "password": password,
                "job_id": jid,
                "job_title": job.title,
                "status": "created"
            })
        except Exception as e:
            errors.append({"row": row_num, "email": email, "errors": [str(e)]})

    return {
        "total_rows": ws.max_row - 1,
        "created": len(results),
        "failed": len(errors),
        "candidates": results,
        "errors": errors
    }
